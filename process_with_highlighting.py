"""
Process Pole Data with Visual Highlighting
===========================================

This script processes pole transfer data and creates an Excel file
with the newly created columns (Marker_Name, Engine_Number, Pole_Number)
highlighted in yellow background for easy identification.

Usage:
    python process_with_highlighting.py input.csv output.xlsx
    python process_with_highlighting.py input.xlsx output.xlsx --sheet "Sheet1"

Author: Max's Automation Team
Date: 2025-12-03
"""

import sys
import argparse
import pandas as pd
from pathlib import Path
from advanced_column_splitter import ColumnSplitter
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def create_highlighted_excel(df: pd.DataFrame, output_path: Path,
                             new_columns: list, original_column: str = None):
    """
    Create an Excel file with specified columns highlighted in yellow.

    Args:
        df: DataFrame to write
        output_path: Path to output Excel file
        new_columns: List of column names to highlight
        original_column: Original raw column name (if kept)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        logger.error("openpyxl is required for Excel highlighting. Install with: pip install openpyxl")
        raise

    logger.info(f"Creating Excel file with highlighting: {output_path}")

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed Data"

    # Define styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # Write data to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border

            # Header row
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            # Data rows - highlight new columns
            elif r_idx > 1:
                col_name = df.columns[c_idx - 1]
                if col_name in new_columns:
                    cell.fill = yellow_fill

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Save workbook
    wb.save(output_path)
    logger.info(f"Successfully created Excel file with {len(df)} rows")


def process_with_visual_output(input_path: str,
                               output_path: str,
                               column_name: str = None,
                               sheet_name: str = None,
                               keep_original: bool = False,
                               remove_duplicates: bool = True,
                               filter_job_numbers: bool = True):
    """
    Process pole data and create visually highlighted Excel output.

    Args:
        input_path: Path to input file (CSV or Excel)
        output_path: Path to output Excel file
        column_name: Name of raw marker column (auto-detected if None)
        sheet_name: Sheet name for Excel input
        keep_original: Keep original raw column
        remove_duplicates: Remove duplicate pole numbers
        filter_job_numbers: Filter out job numbers

    Returns:
        Processing report dictionary
    """
    input_path = Path(input_path)
    output_path = Path(output_path)

    # Ensure output is Excel
    if output_path.suffix.lower() not in ['.xlsx', '.xls']:
        logger.warning(f"Changing output extension to .xlsx for highlighting support")
        output_path = output_path.with_suffix('.xlsx')

    # Validate input
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Read input file
    logger.info(f"Reading input file: {input_path}")
    suffix = input_path.suffix.lower()

    if suffix == '.csv':
        df = pd.read_csv(input_path)
    elif suffix in ['.xlsx', '.xls']:
        if sheet_name:
            df = pd.read_excel(input_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(input_path)
    else:
        raise ValueError(f"Unsupported input format: {suffix}")

    logger.info(f"Loaded {len(df)} rows, {len(df.columns)} columns")

    # Auto-detect column if not specified
    if column_name is None:
        possible_names = [
            'Raw_Marker_Data',
            'Area Section Marker / Installation plan',
            'Marker Data',
            'Marker',
        ]
        for name in possible_names:
            if name in df.columns:
                column_name = name
                logger.info(f"Auto-detected column: '{column_name}'")
                break

        if column_name is None:
            # Try partial match
            for col in df.columns:
                if 'marker' in col.lower() or 'installation' in col.lower():
                    column_name = col
                    logger.info(f"Auto-detected column by keyword: '{column_name}'")
                    break

        if column_name is None:
            raise ValueError("Could not auto-detect column. Use --column parameter.")

    # Process the data
    logger.info("Processing column splitting...")
    splitter = ColumnSplitter(raw_column_name=column_name)

    original_col_name = column_name
    processed_df = splitter.process_dataframe(
        df,
        remove_original=not keep_original,
        filter_job_numbers=filter_job_numbers
    )

    # Remove duplicates if requested
    if remove_duplicates:
        logger.info("Removing duplicate pole numbers...")
        processed_df = splitter.remove_duplicates_by_pole(processed_df)

    # Define which columns to highlight (the new ones)
    new_columns = ['Marker_Name', 'Engine_Number', 'Pole_Number']

    # Create highlighted Excel output
    create_highlighted_excel(
        processed_df,
        output_path,
        new_columns,
        original_col_name if keep_original else None
    )

    # Generate report
    report = splitter.generate_report(processed_df)
    report['input_file'] = str(input_path)
    report['output_file'] = str(output_path)
    report['input_rows'] = len(df)
    report['output_rows'] = len(processed_df)

    return report


def print_report(report: dict):
    """Print formatted report."""
    print("\n" + "=" * 80)
    print("PROCESSING COMPLETE - NEW COLUMNS HIGHLIGHTED IN YELLOW")
    print("=" * 80)
    print(f"\nInput File:  {report['input_file']}")
    print(f"Output File: {report['output_file']}")
    print(f"\n📊 Rows: {report['input_rows']} → {report['output_rows']}")

    if report['input_rows'] != report['output_rows']:
        print(f"   Filtered: {report['input_rows'] - report['output_rows']} rows")

    print(f"\n✅ Successfully Parsed:")
    print(f"   - Engine Numbers: {report['rows_with_engine_number']}")
    print(f"   - Marker Names:   {report['rows_with_marker_name']}")
    print(f"   - Pole Numbers:   {report['rows_with_pole_number']}")

    if report['unparsed_rows'] > 0:
        print(f"\n⚠️  Unparsed Rows: {report['unparsed_rows']} (e.g., Plant Repair)")

    print(f"\n🎨 The following NEW columns are highlighted in YELLOW:")
    print(f"   - Marker_Name")
    print(f"   - Engine_Number")
    print(f"   - Pole_Number")

    print("\n" + "=" * 80)
    print("✨ Open the Excel file to see the highlighted columns!")
    print("=" * 80 + "\n")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Process pole data with visual column highlighting',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s input.csv output.xlsx
  %(prog)s data.xlsx processed.xlsx --sheet "Sheet1"
  %(prog)s input.csv output.xlsx --column "Raw_Marker_Data"
  %(prog)s input.csv output.xlsx --keep-original
        """
    )

    parser.add_argument('input', help='Input file (CSV or Excel)')
    parser.add_argument('output', help='Output Excel file (.xlsx)')
    parser.add_argument('--column', help='Raw marker column name (auto-detected if omitted)')
    parser.add_argument('--sheet', help='Sheet name for Excel input')
    parser.add_argument('--keep-original', action='store_true',
                       help='Keep original raw column')
    parser.add_argument('--no-dedupe', action='store_true',
                       help='Do not remove duplicate pole numbers')
    parser.add_argument('--keep-job-numbers', action='store_true',
                       help='Keep rows with job numbers (JB...)')

    args = parser.parse_args()

    try:
        report = process_with_visual_output(
            input_path=args.input,
            output_path=args.output,
            column_name=args.column,
            sheet_name=args.sheet,
            keep_original=args.keep_original,
            remove_duplicates=not args.no_dedupe,
            filter_job_numbers=not args.keep_job_numbers
        )

        print_report(report)
        return 0

    except Exception as e:
        logger.error(f"Processing failed: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
